from __future__ import annotations

from dataclasses import dataclass, field
from datetime import UTC, datetime
import logging
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from rfp_rag_assistant.loaders import LoadedDocument
from rfp_rag_assistant.models import ParsedDocument, ParsedSection
from rfp_rag_assistant.parsers.title_normalization import normalize_section_title


@dataclass(slots=True)
class ResponseSupportingMaterialExcelParser:
    document_type: str = "response_supporting_material"
    header_scan_rows: int = 20
    logger: logging.Logger = field(default_factory=lambda: logging.getLogger(__name__))

    def parse(self, document: LoadedDocument) -> ParsedDocument:
        if isinstance(document.payload, (str, Path)):
            return self.parse_file(Path(document.payload))
        return self.parse_file(document.source_file)

    def parse_file(self, source_file: Path) -> ParsedDocument:
        workbook = load_workbook(source_file, data_only=True)

        sections: list[ParsedSection] = []
        sheet_summaries: list[dict[str, Any]] = []
        row_section_count = 0
        row_group_section_count = 0
        summary_section_count = 0

        for worksheet in workbook.worksheets:
            sheet_sections = self._parse_sheet(source_file, worksheet)
            sections.extend(sheet_sections)
            row_section_count += sum(1 for section in sheet_sections if section.kind == "spreadsheet_row")
            row_group_section_count += sum(1 for section in sheet_sections if section.kind == "spreadsheet_row_group")
            summary_section_count += sum(1 for section in sheet_sections if section.kind == "reference_content")
            sheet_summaries.append(
                {
                    "sheet_name": worksheet.title,
                    "section_count": len(sheet_sections),
                }
            )
        self.logger.info(
            "Parsed response supporting material workbook=%s sheets=%s sections=%s row_sections=%s row_group_sections=%s summary_sections=%s",
            source_file.name,
            len(workbook.sheetnames),
            len(sections),
            row_section_count,
            row_group_section_count,
            summary_section_count,
        )

        return ParsedDocument(
            source_file=source_file,
            file_type=source_file.suffix.lstrip(".").lower(),
            document_type=self.document_type,
            extracted_at=datetime.now(UTC),
            sections=sections,
            metadata={
                "subtype": "excel_supporting_material",
                "sheet_count": len(workbook.sheetnames),
                "sheet_summaries": sheet_summaries,
            },
        )

    def _parse_sheet(self, source_file: Path, worksheet: Any) -> list[ParsedSection]:
        rows = [[self._clean_value(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
        non_empty_rows = [
            (row_index, row)
            for row_index, row in enumerate(rows, start=1)
            if any(cell for cell in row)
        ]
        if not non_empty_rows:
            return []

        profile_sections = self._parse_contact_profile_sheet(source_file, worksheet.title, non_empty_rows)
        if profile_sections:
            return profile_sections

        header_row_index, headers = self._detect_headers(non_empty_rows)
        if header_row_index is None or not headers:
            return [self._build_sheet_summary_section(source_file, worksheet.title, non_empty_rows)]

        sections: list[ParsedSection] = []
        preamble_rows = [row for row_index, row in non_empty_rows if row_index < header_row_index]
        sheet_context = self._render_sheet_context(preamble_rows)

        for row_index, row in non_empty_rows:
            if row_index <= header_row_index:
                continue
            record = self._row_record(headers, row)
            if not record:
                continue
            section_id = self._slugify(f"{worksheet.title}-row-{row_index}")
            sections.append(
                ParsedSection(
                    section_id=section_id,
                    title=f"{worksheet.title} row {row_index}",
                    text=self._render_row_text(
                        source_file=source_file,
                        sheet_name=worksheet.title,
                        row_index=row_index,
                        record=record,
                        sheet_context=sheet_context,
                    ),
                    kind="spreadsheet_row",
                    heading_path=(worksheet.title,),
                    structured_data={
                        "sheet_name": worksheet.title,
                        "row_index": row_index,
                        "record": record,
                        "sheet_context": sheet_context,
                        "section_title_normalized": normalize_section_title(f"{worksheet.title} row {row_index}"),
                    },
                )
            )

        if sections:
            return sections
        return [self._build_sheet_summary_section(source_file, worksheet.title, non_empty_rows)]

    def _detect_headers(self, non_empty_rows: list[tuple[int, list[str]]]) -> tuple[int | None, list[str]]:
        candidates = [
            (row_index, row)
            for row_index, row in non_empty_rows
            if row_index <= self.header_scan_rows
        ]
        if not candidates:
            return None, []

        best_index: int | None = None
        best_headers: list[str] = []
        best_score = 0

        for row_index, row in candidates:
            cleaned = [cell.strip() for cell in row if cell.strip()]
            score = len(cleaned)
            if score < 2:
                continue
            keyword_bonus = sum(3 for cell in cleaned if self._looks_like_header_keyword(cell))
            score += keyword_bonus
            if any(self._looks_like_data_value(cell) for cell in cleaned[: min(3, len(cleaned))]) and keyword_bonus == 0:
                continue
            if all(self._looks_like_data_value(cell) for cell in cleaned[: min(2, len(cleaned))]):
                continue
            if score > best_score:
                best_index = row_index
                best_headers = self._normalise_headers(row)
                best_score = score

        return best_index, best_headers

    def _parse_contact_profile_sheet(
        self,
        source_file: Path,
        sheet_name: str,
        non_empty_rows: list[tuple[int, list[str]]],
    ) -> list[ParsedSection]:
        if not self._looks_like_contact_profile_sheet(non_empty_rows):
            return []

        max_columns = max(len(row) for _, row in non_empty_rows)
        sections: list[ParsedSection] = []
        for column_index in range(max_columns):
            profile_fields: dict[str, str] = {}
            narrative_parts: list[str] = []
            for _, row in non_empty_rows:
                if column_index >= len(row):
                    continue
                cell = row[column_index].strip()
                if not cell:
                    continue
                if ":" in cell:
                    label, value = cell.split(":", 1)
                    label = label.strip().lower()
                    value = value.strip()
                    if label in {"name", "title", "email", "phone"} and value:
                        if label == "name" and "name" in profile_fields:
                            sections.append(
                                self._build_profile_section(
                                    source_file=source_file,
                                    sheet_name=sheet_name,
                                    profile_fields=profile_fields,
                                    narrative_parts=narrative_parts,
                                )
                            )
                            profile_fields = {}
                            narrative_parts = []
                        profile_fields[label] = value
                        continue
                if not self._looks_like_header_keyword(cell):
                    narrative_parts.append(cell)

            if "name" not in profile_fields:
                continue

            sections.append(
                self._build_profile_section(
                    source_file=source_file,
                    sheet_name=sheet_name,
                    profile_fields=profile_fields,
                    narrative_parts=narrative_parts,
                )
            )

        return sections

    def _build_profile_section(
        self,
        *,
        source_file: Path,
        sheet_name: str,
        profile_fields: dict[str, str],
        narrative_parts: list[str],
    ) -> ParsedSection:
        title = f"{sheet_name} profile - {profile_fields['name']}"
        return ParsedSection(
            section_id=self._slugify(title),
            title=title,
            text=self._render_profile_text(
                source_file=source_file,
                sheet_name=sheet_name,
                profile_fields=profile_fields,
                narrative_parts=narrative_parts,
            ),
            kind="spreadsheet_row_group",
            heading_path=(sheet_name,),
            structured_data={
                "sheet_name": sheet_name,
                "profile_fields": profile_fields,
                "section_title_normalized": normalize_section_title(title),
            },
        )

    def _normalise_headers(self, row: list[str]) -> list[str]:
        headers: list[str] = []
        for index, value in enumerate(row, start=1):
            header = value.strip()
            if not header:
                header = f"column_{index}"
            headers.append(header)
        return headers

    def _row_record(self, headers: list[str], row: list[str]) -> dict[str, str]:
        values = [value.strip() for value in row]
        record: dict[str, str] = {}
        for index, header in enumerate(headers):
            if index >= len(values):
                break
            value = values[index]
            if value:
                record[header] = value
        return record

    def _render_row_text(
        self,
        *,
        source_file: Path,
        sheet_name: str,
        row_index: int,
        record: dict[str, str],
        sheet_context: str,
    ) -> str:
        parts = [
            (
                f"This row from sheet '{sheet_name}' in file '{source_file.name}' "
                f"captures response supporting material at row {row_index}."
            )
        ]
        if sheet_context:
            parts.append(f"Sheet context: {sheet_context}")
        parts.append("Row values: " + " | ".join(f"{key}: {value}" for key, value in record.items()))
        return "\n\n".join(parts).strip()

    def _render_profile_text(
        self,
        *,
        source_file: Path,
        sheet_name: str,
        profile_fields: dict[str, str],
        narrative_parts: list[str],
    ) -> str:
        parts = [
            (
                f"This profile from sheet '{sheet_name}' in file '{source_file.name}' "
                f"captures authored supporting material for the role '{profile_fields.get('title', profile_fields['name'])}'."
            )
        ]
        ordered_fields = ["name", "title", "email", "phone"]
        field_text = " | ".join(
            f"{field.title()}: {profile_fields[field]}"
            for field in ordered_fields
            if field in profile_fields
        )
        if field_text:
            parts.append(field_text)
        if narrative_parts:
            parts.append("Responsibilities: " + " ".join(narrative_parts[:3]))
        return "\n\n".join(parts).strip()

    def _render_sheet_context(self, rows: list[list[str]]) -> str:
        parts: list[str] = []
        for row in rows:
            text = " | ".join(value for value in row if value)
            if text:
                parts.append(text)
        return " ".join(parts[:5]).strip()

    def _build_sheet_summary_section(
        self,
        source_file: Path,
        sheet_name: str,
        non_empty_rows: list[tuple[int, list[str]]],
    ) -> ParsedSection:
        content_rows = [" | ".join(value for value in row if value) for _, row in non_empty_rows[:20]]
        text = (
            f"Sheet '{sheet_name}' from file '{source_file.name}' contains supporting material.\n\n"
            + "\n".join(row for row in content_rows if row)
        ).strip()
        return ParsedSection(
            section_id=self._slugify(f"{sheet_name}-summary"),
            title=f"{sheet_name} summary",
            text=text,
            kind="reference_content",
            heading_path=(sheet_name,),
            structured_data={
                "sheet_name": sheet_name,
                "summary_only": True,
                "section_title_normalized": normalize_section_title(f"{sheet_name} summary"),
            },
        )

    def _clean_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        return str(value).strip()

    def _looks_like_data_value(self, value: str) -> bool:
        compact = value.strip()
        if not compact:
            return False
        if re.fullmatch(r"\d+([./-]\d+)*", compact):
            return True
        if "@" in compact:
            return True
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", compact):
            return True
        return False

    def _looks_like_header_keyword(self, value: str) -> bool:
        compact = value.strip().lower().rstrip(":")
        header_terms = {
            "task no",
            "description",
            "status",
            "owner",
            "start date",
            "end date",
            "task completion",
            "name",
            "title",
            "email",
            "phone",
            "authority & other",
            "blatchford",
            "nwd",
            "dur",
        }
        return compact in header_terms

    def _looks_like_contact_profile_sheet(self, non_empty_rows: list[tuple[int, list[str]]]) -> bool:
        sample_cells = [
            cell.strip()
            for row_index, row in non_empty_rows
            if row_index <= 20
            for cell in row
            if cell.strip()
        ]
        field_hits = sum(
            1
            for cell in sample_cells
            if any(cell.lower().startswith(prefix) for prefix in ("name:", "title:", "email:", "phone:"))
        )
        return field_hits >= 6

    def _slugify(self, value: str) -> str:
        slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
        return slug or "sheet"
